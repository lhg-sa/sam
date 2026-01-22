from pathlib import Path

import frappe
from openpyxl import load_workbook


DOCTYPE_FIELDS = {
    "codigo_insumo",
    "nombre_insumo",
    "renglon_presupuestario",
    "presentacion",
    "udm_insumo",
    "caracteristicas",
    "codigo_presentacion",
    "es_activo_fijo",
    "clase",
}

REQUIRED_FIELDS = {"codigo_insumo", "nombre_insumo", "renglon_presupuestario"}

NAMING_PATTERN = "format:{codigo_insumo}-{renglon_presupuestario}"


def _sanitize_utf8(value):
    if value is None:
        return value
    if not isinstance(value, str):
        return value
    value = value.replace("\x00", "").replace("\ufffd", "")
    value = value.encode("utf-8", errors="replace").decode("utf-8")
    return value.replace("\ufffd", "")


def _normalize_cell(value):
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        return _sanitize_utf8(value.strip())
    return value


def import_dafim_catalogo_insumos(limit=None):
    app_path = Path(frappe.get_app_path("sam"))
    file_path = app_path / "utils" / "insumos_excel.xlsx"
    log_path = app_path / "utils" / "import_dafim_catalogo_insumos_errors.log"
    doctype = "DAFIM Catalogo Insumos"

    if not file_path.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {file_path}")

    meta = frappe.get_meta(doctype)
    if meta.autoname != NAMING_PATTERN:
        raise ValueError(
            f"El autoname del DocType ({meta.autoname}) no coincide con el esperado ({NAMING_PATTERN})."
        )

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)

    try:
        header_row = next(rows)
    except StopIteration:
        raise ValueError("El archivo Excel esta vacio.")

    column_indexes = {}
    for idx, header in enumerate(header_row):
        if not isinstance(header, str):
            continue
        cleaned = header.strip().lower()
        if cleaned and cleaned in DOCTYPE_FIELDS:
            column_indexes[cleaned] = idx

    missing_columns = [col for col in REQUIRED_FIELDS if col not in column_indexes]
    if missing_columns:
        raise ValueError(
            f"El archivo Excel no contiene las columnas requeridas: {', '.join(missing_columns)}"
        )

    inserted = 0
    seen_combinations = set()
    for excel_row_index, row in enumerate(rows, start=2):
        if not any(cell not in (None, "") for cell in row):
            continue

        if limit is not None and inserted >= limit:
            break

        values = {}
        for fieldname in DOCTYPE_FIELDS:
            column_index = column_indexes.get(fieldname)
            raw_value = row[column_index] if column_index is not None else None
            values[fieldname] = _normalize_cell(raw_value)

        mapped_indexes = set(column_indexes.values())
        extras = []
        for index, cell in enumerate(row):
            if index in mapped_indexes:
                continue
            cleaned = _normalize_cell(cell)
            if cleaned not in (None, ""):
                extras.append(str(cleaned))

        if extras:
            if values.get("caracteristicas"):
                values["caracteristicas"] = f"{values['caracteristicas']} {' '.join(extras)}"
            else:
                values["caracteristicas"] = " ".join(extras)

        codigo_insumo = values.get("codigo_insumo")
        renglon_presupuestario = values.get("renglon_presupuestario")

        if not codigo_insumo or not values.get("nombre_insumo") or not renglon_presupuestario:
            error_line = (
                f"Fila {excel_row_index} omitida (datos requeridos faltantes). "
                f"Valores: codigo_insumo={codigo_insumo}, nombre_insumo={values.get('nombre_insumo')}, "
                f"renglon_presupuestario={renglon_presupuestario}\n"
            )
            with log_path.open("a", encoding="utf-8") as logfile:
                logfile.write(error_line)
            continue

        combination_key = (codigo_insumo, renglon_presupuestario)
        if combination_key in seen_combinations:
            continue
        seen_combinations.add(combination_key)

        if frappe.db.exists(
            doctype,
            {"codigo_insumo": codigo_insumo, "renglon_presupuestario": renglon_presupuestario},
        ):
            print(
                f"Fila {excel_row_index} omitida "
                f"(codigo_insumo y renglon_presupuestario ya existen): "
                f"{codigo_insumo} / {renglon_presupuestario}"
            )
            continue

        try:
            doc = frappe.new_doc(doctype)

            for fieldname, value in values.items():
                if fieldname in {"nombre_insumo", "caracteristicas"} and isinstance(value, str):
                    value = value.strip()
                if fieldname == "nombre_insumo" and isinstance(value, str):
                    value = value[:140]
                doc.set(fieldname, value)

            doc.insert(ignore_permissions=True)
            inserted += 1
            print(f"Fila {excel_row_index} insertada correctamente")
        except Exception as exc:
            error_line = (
                f"Fila {excel_row_index} (codigo_insumo={codigo_insumo}): {exc}\n"
            )
            with log_path.open("a", encoding="utf-8") as logfile:
                logfile.write(error_line)

    frappe.db.commit()
    print("Importacion completada.")
